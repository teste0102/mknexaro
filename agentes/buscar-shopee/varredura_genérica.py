#!/usr/bin/env python3
"""
Wrapper genérico para varreduras de preços
Suporta: Shopee, Mercado Livre, qualquer marketplace
Uso: python3 varredura_genérica.py <script> "termo"
"""

import subprocess
import json
import re
import sys
import os
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class VarreduraGenérica:
    def __init__(self, script_path, termo):
        self.script_path = script_path
        self.termo = termo
        self.resultados = []
        self.nome_marketplace = Path(script_path).stem.upper()
    
    def executar_script(self):
        """Executa o script de scraping"""
        try:
            print(f"🔍 Executando {self.nome_marketplace}...")
            resultado = subprocess.run(
                ['python3', self.script_path, self.termo],
                capture_output=True,
                text=True,
                timeout=120
            )
            
            if resultado.returncode != 0:
                print(f"⚠️  Script retornou erro: {resultado.stderr}")
                return None
            
            return resultado.stdout
        except Exception as e:
            print(f"❌ Erro ao executar script: {e}")
            return None
    
    def parsear_output(self, output):
        """Parseia output automaticamente (HTML, JSON, etc)"""
        
        # Tenta JSON primeiro
        try:
            dados = json.loads(output)
            if isinstance(dados, list):
                return dados
            return [dados]
        except:
            pass
        
        # Tenta HTML
        if '<html' in output.lower() or '<div' in output.lower():
            soup = BeautifulSoup(output, 'html.parser')
            precos = re.findall(r'R\$\s*([\d.,]+)', output)
            produtos = re.findall(r'<h\d[^>]*>([^<]+)</h\d>', output)
            
            return [{
                'produto': p[:50],
                'preco': f"R$ {preco}",
                'marketplace': self.nome_marketplace
            } for p, preco in zip(produtos, precos)]
        
        # Tenta extrair preços com regex direto
        precos = re.findall(r'R\$\s*([\d.,]+)', output)
        if precos:
            return [{
                'produto': f'Produto {i+1}',
                'preco': f"R$ {preco}",
                'marketplace': self.nome_marketplace
            } for i, preco in enumerate(precos[:5])]
        
        return []
    
    def processar(self):
        """Pipeline completo: executar → parsear → normalizar"""
        output = self.executar_script()
        if not output:
            return []
        
        self.resultados = self.parsear_output(output)
        
        # Normaliza estrutura
        for r in self.resultados:
            if 'marketplace' not in r:
                r['marketplace'] = self.nome_marketplace
        
        return self.resultados

def salvar_excel(todas_varreduras, termo):
    """Salva múltiplas varreduras em um Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Preços"
    
    # Header
    headers = ['Marketplace', 'Produto', 'Preço', 'Data']
    ws.append(headers)
    
    # Formatação
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Adiciona dados
    data_hoje = datetime.now().strftime('%d/%m/%Y')
    for marketplace_dados in todas_varreduras:
        for row in marketplace_dados:
            ws.append([
                row.get('marketplace', 'N/A'),
                row.get('produto', 'N/A'),
                row.get('preco', 'N/A'),
                data_hoje
            ])
    
    # Ajusta colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    filename = f"precos_{termo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python3 varredura_genérica.py <script_shopee> <script_ml> 'termo'")
        print("Ex:  python3 varredura_genérica.py scraper.py ../buscar-ml/ml_scraper.py 'tela A15'")
        sys.exit(1)
    
    # Suporta múltiplos scripts
    scripts = sys.argv[1:-1]
    termo = sys.argv[-1]
    
    print(f"\n🚀 VARREDURA GENÉRICA - Buscando: '{termo}'\n")
    
    todos_resultados = []
    
    for script_path in scripts:
        if not os.path.exists(script_path):
            print(f"⚠️  Script não encontrado: {script_path}")
            continue
        
        varredura = VarreduraGenérica(script_path, termo)
        resultados = varredura.processar()
        
        if resultados:
            print(f"✅ {varredura.nome_marketplace}: {len(resultados)} produtos encontrados")
            todos_resultados.append(resultados)
        else:
            print(f"❌ {varredura.nome_marketplace}: sem resultados")
    
    if todos_resultados:
        arquivo = salvar_excel(todos_resultados, termo)
        print(f"\n📊 Excel gerado: {arquivo}")
    else:
        print("\n❌ Nenhum resultado encontrado em nenhum marketplace")
