#!/usr/bin/env python3
"""
Script para criar Excel a partir de dados coletados manualmente da Shopee
Uso: python3 criar_excel_coleta.py "termo de busca"
"""
import sys
import json
import os
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def criar_excel_vazio(termo):
    """Cria um template Excel vazio para coleta manual"""

    # Criar DataFrame vazio com as colunas corretas
    colunas = ['Título', 'Preço', 'Desconto', 'Avaliação', 'Vendas', 'URL', 'Fonte', 'Pasta', 'Imagem']
    df = pd.DataFrame(columns=colunas)

    # Criar pasta de saída
    pasta_saida = Path.home() / 'dados_produtos' / 'shopee'
    pasta_saida.mkdir(parents=True, exist_ok=True)

    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_arquivo = f"{timestamp}_{termo.replace(' ', '_')}.xlsx"
    caminho = pasta_saida / nome_arquivo

    # Salvar Excel
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Produtos', index=False)
        ws = writer.sheets['Produtos']

        # Formatação
        header_font = Font(name='Calibri', bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar formatação ao cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar largura das colunas
        col_widths = {
            'A': 60,  # Título
            'B': 15,  # Preço
            'C': 12,  # Desconto
            'D': 12,  # Avaliação
            'E': 15,  # Vendas
            'F': 30,  # URL
            'G': 20,  # Fonte
            'H': 15,  # Pasta
            'I': 20,  # Imagem
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Adicionar algumas linhas vazias para preenchimento
        for row in range(2, 22):
            for cell in ws[row]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

    return str(caminho)

def converter_json_para_excel(arquivo_json):
    """Converte JSON coletado do console para Excel"""
    try:
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            produtos = json.load(f)

        # Converter para DataFrame
        df = pd.DataFrame(produtos)

        # Garantir que tem todas as colunas
        colunas_esperadas = ['Título', 'Preço', 'Desconto', 'Avaliação', 'Vendas', 'URL', 'Fonte', 'Pasta', 'Imagem']
        for col in colunas_esperadas:
            if col not in df.columns:
                df[col] = ''

        # Salvar
        pasta_saida = Path.home() / 'dados_produtos' / 'shopee'
        pasta_saida.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f"{timestamp}_coletados.xlsx"
        caminho = pasta_saida / nome_arquivo

        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df[colunas_esperadas].to_excel(writer, sheet_name='Produtos', index=False)

        return str(caminho)

    except Exception as e:
        print(f"Erro: {e}")
        return None

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 criar_excel_coleta.py 'termo de busca'")
        sys.exit(1)

    termo = sys.argv[1]

    print(f"📊 Criando template Excel para: {termo}")
    caminho = criar_excel_vazio(termo)

    print(f"\n✅ Excel criado em:")
    print(f"   {caminho}")
    print(f"\n📝 Próximos passos:")
    print(f"   1. Abra o arquivo Excel")
    print(f"   2. Cole os dados coletados do console da Shopee")
    print(f"   3. Salve o arquivo")
    print(f"   4. Use comparador.py para comparar preços")
