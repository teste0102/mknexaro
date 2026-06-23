import os
import json
import re
import csv
from bs4 import BeautifulSoup
from datetime import datetime
import logging
from collections import Counter
from urllib.parse import urljoin, urlparse, unquote
from pathlib import Path

# Configuração de logging melhorada
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('shopee_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

driver = None
parar_execucao = False


class ShopeeScraper:
    """Scraper especializado para produtos da Shopee com saída melhorada"""
    
    def __init__(self, caminho_arquivo: str):
        self.caminho_arquivo = caminho_arquivo
        self.soup = None
        self.produtos = []
        self.estatisticas = {}
        self.base_url = "https://shopee.com.br"
        
    def carregar_html(self) -> bool:
        """Carrega o arquivo HTML"""
        try:
            if not os.path.exists(self.caminho_arquivo):
                logger.error(f"Arquivo não encontrado: {self.caminho_arquivo}")
                return False
                
            with open(self.caminho_arquivo, 'r', encoding='utf-8') as f:
                html = f.read()
                
            self.soup = BeautifulSoup(html, 'html.parser')
            logger.info("HTML carregado com sucesso")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao carregar HTML: {e}")
            return False
    
    def analisar_estrutura(self):
        """Analisa a estrutura do HTML para identificar padrões de produtos"""
        if not self.soup:
            return
            
        print("\n" + "="*60)
        print("          ANÁLISE DE ESTRUTURA DO HTML")
        print("="*60)
        
        # Estatísticas gerais
        total_divs = len(self.soup.find_all('div'))
        total_links = len(self.soup.find_all('a'))
        total_imagens = len(self.soup.find_all('img'))
        
        print(f"📊 Estatísticas Gerais:")
        print(f"   • Total de DIVs: {total_divs:,}")
        print(f"   • Total de Links: {total_links:,}")
        print(f"   • Total de Imagens: {total_imagens:,}")
        
        # Busca por padrões comuns em sites de e-commerce
        print(f"\n🔍 Analisando containers de produtos...")
        
        # Busca por elementos com preços (R$)
        elementos_preco = self.soup.find_all(string=re.compile(r'R\$\s*[\d.,]+'))
        print(f"💰 Elementos com preço encontrados: {len(elementos_preco)}")
        
        if elementos_preco:
            # Analisa os preços encontrados
            precos_limpos = []
            for preco in elementos_preco:
                match = re.search(r'R\$\s*([\d.,]+)', preco)
                if match:
                    valor_str = match.group(1).replace('.', '').replace(',', '.')
                    try:
                        valor = float(valor_str)
                        precos_limpos.append(valor)
                    except:
                        pass
            
            if precos_limpos:
                print(f"   • Menor preço: R$ {min(precos_limpos):.2f}")
                print(f"   • Maior preço: R$ {max(precos_limpos):.2f}")
                print(f"   • Preço médio: R$ {sum(precos_limpos)/len(precos_limpos):.2f}")
        
        # Mostra classes mais comuns
        classes_counter = Counter()
        for elem in self.soup.find_all(class_=True):
            for classe in elem.get('class', []):
                classes_counter[classe] += 1
        
        print(f"\n📦 Classes CSS mais comuns:")
        for classe, count in classes_counter.most_common(10):
            print(f"   • {classe}: {count} ocorrências")
    
    def extrair_produtos_shopee(self):
        """Extrai produtos usando seletores específicos da Shopee"""
        if not self.soup:
            logger.error("HTML não carregado")
            return []

        produtos = []
        
        # Tenta extrair usando diferentes seletores
        seletores = [
            'div[data-sqe="item"]',
            'div.shopee-search-item-result__item',
            '.col-xs-2-4'
        ]

        # Tenta cada seletor até encontrar produtos
        for seletor in seletores:
            if parar_execucao:
                print("[INFO] Execução interrompida pelo usuário!")
                break
            elementos = self.soup.select(seletor)
            if elementos:
                print(f"Encontrados {len(elementos)} elementos com o seletor: {seletor}")
                for elemento in elementos:
                    produto = {}
                    
                    # Extrai título e URL
                    link = elemento.select_one('a[href]')
                    if link:
                        href = link.get('href')
                        if href:
                            produto['url'] = urljoin(self.base_url, href) if href.startswith('/') else href
                        
                        # Tenta extrair título
                        titulo = link.get('title', '') or link.get_text(strip=True)
                        if titulo:
                            produto['titulo'] = titulo

                    # Extrai imagem
                    img = elemento.select_one('img')
                    if img:
                        src = img.get('src') or img.get('data-src')
                        if src:
                            produto['imagem'] = src

                    # Extrai preço
                    preco = elemento.select_one('[class*="price"]')
                    if preco:
                        produto['preco'] = preco.get_text(strip=True)

                    # Extrai avaliação
                    rating = elemento.select_one('[class*="rating"]')
                    if rating:
                        produto['avaliacao'] = rating.get_text(strip=True)

                    # Extrai vendas
                    vendas = elemento.select_one('[class*="sold"]')
                    if vendas:
                        produto['vendas'] = vendas.get_text(strip=True)

                    # Extrai desconto
                    desconto = elemento.select_one('[class*="percent"]')
                    if desconto:
                        produto['desconto'] = desconto.get_text(strip=True)

                    if produto:
                        produtos.append(produto)

                if produtos:
                    break  # Se encontrou produtos, não precisa tentar outros seletores

        # Limpa e refina os produtos
        produtos = self._limpar_produtos(produtos)
        
        # Calcula estatísticas
        self.produtos = produtos
        self._calcular_estatisticas()

        return produtos
    
    def _extrair_por_links(self):
        """Extrai produtos através de links específicos da Shopee"""
        produtos = []
        
        # Padrões de URL da Shopee
        padroes_url = [
            r'shopee\.com\.br/.*-i\.\d+\.\d+',
            r'/.*-i\.\d+\.\d+',
            r'product/\d+',
        ]
        
        links_encontrados = set()
        
        for padrao in padroes_url:
            links = self.soup.find_all('a', href=re.compile(padrao))
            
            for link in links:
                href = link.get('href')
                if href not in links_encontrados:
                    links_encontrados.add(href)
                    produto = self._extrair_info_completa_do_link(link)
                    if produto and produto.get('titulo'):
                        produtos.append(produto)
        
        return produtos
    
    def _extrair_info_completa_do_link(self, link):
        """Extrai informações completas a partir de um link de produto"""
        produto = {}
        
        # URL do produto
        href = link.get('href', '')
        if href.startswith('/'):
            href = urljoin(self.base_url, href)
        produto['url'] = href
        
        # Tenta extrair o título do atributo title do link primeiro
        if link.get('title'):
            titulo = link.get('title').strip()
            # Verifica se não é vendas ou porcentagem
            if (titulo and 
                not re.match(r'^\d+\s*(vendido|vendidos)', titulo) and
                not re.match(r'^-?\d+%', titulo)):
                produto['titulo'] = titulo
        
        # Se não tem título, extrai do texto do link
        if not produto.get('titulo'):
            texto_link = link.get_text(strip=True)
            if (texto_link and len(texto_link) > 10 and
                not re.match(r'^\d+\s*(vendido|vendidos)', texto_link) and
                not texto_link.startswith('R$')):
                produto['titulo'] = texto_link
        
        # Busca no container pai
        container = link.parent
        niveis = 0
        
        while container and niveis < 4:
            # Busca título se ainda não encontrou
            if not produto.get('titulo'):
                # Busca em elementos com atributo title
                titulo_elem = container.find(['div', 'span'], attrs={'title': True})
                if titulo_elem:
                    titulo = titulo_elem.get('title')
                    if (titulo and len(titulo) > 10 and
                        not re.match(r'^\d+\s*(vendido|vendidos)', titulo) and
                        not re.match(r'^-?\d+%', titulo)):
                        produto['titulo'] = titulo
                        
                # Se não achou, busca por classes específicas
                if not produto.get('titulo'):
                    for classe in ['truncate', 'product-name', 'item-name']:
                        elem = container.find(['div', 'span'], class_=re.compile(classe))
                        if elem:
                            texto = elem.get_text(strip=True)
                            if (texto and len(texto) > 10 and 
                                not re.match(r'^\d+\s*(vendido|vendidos)', texto) and
                                not texto.startswith('R$') and
                                not re.match(r'^-?\d+%', texto)):
                                produto['titulo'] = texto
                                break
            
            # Busca preço
            if not produto.get('preco'):
                # Busca todos os textos com R$
                for texto in container.stripped_strings:
                    if 'R$' in texto:
                        match = re.search(r'R\$\s*([\d.,]+)', texto)
                        if match:
                            produto['preco'] = match.group(0)
                            break
                
                # Busca em elementos específicos
                if not produto.get('preco'):
                    preco_elem = container.find(['div', 'span'], class_=re.compile('price|cost|valor'))
                    if preco_elem:
                        texto = preco_elem.get_text(strip=True)
                        match = re.search(r'R\$\s*([\d.,]+)', texto)
                        if match:
                            produto['preco'] = match.group(0)
            
            # Busca vendas
            if not produto.get('vendas'):
                for texto in container.stripped_strings:
                    match = re.search(r'(\d+(?:\.\d+)?[kK]?)\s*(vendido|vendidos)', texto, re.I)
                    if match:
                        numero = match.group(1)
                        if 'k' in numero.lower():
                            numero = numero.replace('k', '').replace('K', '')
                            try:
                                numero = str(int(float(numero) * 1000))
                            except:
                                pass
                        produto['vendas'] = f"{numero} vendidos"
                        break
            
            # Busca avaliação
            if not produto.get('avaliacao'):
                for texto in container.stripped_strings:
                    if re.match(r'^[0-5]\.[0-9]$', texto.strip()):
                        produto['avaliacao'] = texto.strip()
                        break
            
            # Busca imagem
            if not produto.get('imagem'):
                img = container.find('img')
                if img:
                    src = img.get('data-src') or img.get('src') or img.get('data-original')
                    if src and not src.startswith('data:'):
                        if src.startswith('//'):
                            src = 'https:' + src
                        elif src.startswith('/'):
                            src = urljoin(self.base_url, src)
                        produto['imagem'] = src
            
            # Busca loja
            if not produto.get('loja'):
                loja_elem = container.find(['div', 'span', 'a'], class_=re.compile('shop|seller|store|loja'))
                if loja_elem:
                    loja_texto = loja_elem.get_text(strip=True)
                    if (loja_texto and 
                        not loja_texto.startswith('R$') and 
                        not re.match(r'^-?\d+%', loja_texto) and
                        not re.match(r'^\d+\.\d+$', loja_texto) and
                        len(loja_texto) > 2):
                        produto['loja'] = loja_texto
            
            # Busca desconto
            if not produto.get('desconto'):
                for texto in container.stripped_strings:
                    match = re.search(r'-(\d+)%', texto)
                    if match:
                        produto['desconto'] = match.group(0)
                        break
            
            # Sobe um nível
            container = container.parent
            niveis += 1
        
        # Retorna apenas se tem título ou preço
        return produto if (produto.get('titulo') or produto.get('preco')) else None
    
    def _extrair_por_grid(self):
        """Extrai produtos de estruturas em grid"""
        produtos = []
        
        # Busca por containers de grid comuns
        seletores_grid = [
            'div[class*="grid"]',
            'div[class*="list"]',
            'div[class*="item-result"]',
            'div[class*="product"]',
            '.row > div',
            '[class*="col-"]'
        ]
        
        for seletor in seletores_grid:
            elementos = self.soup.select(seletor)
            
            for elemento in elementos:
                if parar_execucao:
                    print("[INFO] Execução interrompida pelo usuário!")
                    break
                # Verifica se tem características de produto
                tem_preco = bool(elemento.find(string=re.compile(r'R\$')))
                tem_imagem = bool(elemento.find('img'))
                tem_link = bool(elemento.find('a'))
                
                if tem_preco or (tem_imagem and tem_link):
                    produto = self._extrair_info_produto(elemento)
                    if produto:
                        produtos.append(produto)
        
        return produtos
    
    def _extrair_por_css(self):
        """Extrai usando seletores CSS específicos da Shopee"""
        produtos = []
        
        seletores_especificos = [
            'div[data-sqe="item"]',
            'a[data-sqe="link"]',
            '.col-xs-2-4',
            'div.shopee-search-item-result__item',
            '[class*="search-item"]',
            '[class*="item-card"]'
        ]
        
        for seletor in seletores_especificos:
            elementos = self.soup.select(seletor)
            
            for elemento in elementos:
                produto = self._extrair_info_produto(elemento)
                if produto:
                    produtos.append(produto)
        
        return produtos
    
    def _extrair_preco(self, elemento):
        """
        Extrai o preço usando uma abordagem baseada em padrões conhecidos da Shopee
        """
        # Lista de todos os textos do elemento
        textos = [text.strip() for text in elemento.stripped_strings]
        
        # Filtra apenas textos que contêm R$
        precos_candidatos = []
        for texto in textos:
            if 'R$' in texto:
                # Limpa o texto
                texto = texto.replace('.', '')  # Remove pontos de milhar
                
                # Procura padrão R$ XX,XX
                match = re.search(r'R\$\s*(\d+)[,](\d{2})\b', texto)
                if match:
                    reais, centavos = match.groups()
                    try:
                        valor = float(f"{reais}.{centavos}")
                        # Shopee raramente tem produtos abaixo de R$1 ou acima de R$500
                        if 1 <= valor <= 500:
                            precos_candidatos.append((valor, f"R$ {reais},{centavos}"))
                    except:
                        continue
        
        if precos_candidatos:
            # Ordena por valor e pega o primeiro (geralmente o mais provável)
            precos_candidatos.sort(key=lambda x: x[0])
            return precos_candidatos[0][1]
        
        return None

    def _validar_preco(self, preco):
        """
        Valida se um preço parece correto baseado em padrões conhecidos
        """
        if not preco:
            return False
        
        try:
            # Remove R$ e converte para float
            valor = float(preco.replace('R$', '').replace(' ', '').replace(',', '.'))
            
            # Validações específicas da Shopee
            if valor < 1 or valor > 500:  # Faixa de preço improvável
                return False
            if '.' in preco and ',' in preco:  # Não deve ter ponto e vírgula
                return False
            if len(preco.split(',')[1]) != 2:  # Deve ter exatamente 2 decimais
                return False
                
            return True
        except:
            return False

    def _limpar_preco(self, texto):
        """Limpa e formata o preço"""
        try:
            # Remove tudo que não seja números, R$, vírgula ou ponto
            texto = re.sub(r'[^\d.,R$]', '', texto)
            
            # Garante que só tem um R$
            if texto.count('R$') > 1:
                texto = 'R$' + texto.split('R$')[-1]
                
            # Pega apenas o primeiro conjunto de números após R$
            match = re.search(r'R\$\s*(\d+)[.,](\d{2})', texto)
            if match:
                reais, centavos = match.groups()
                # Remove pontos do valor em reais
                reais = reais.replace('.', '')
                # Converte para ter certeza que é um número válido
                valor = float(f"{reais}.{centavos}")
                if 0.5 <= valor <= 500:  # Faixa de preço típica da Shopee
                    return f"R$ {reais},{centavos}"
        except:
            pass
        return None

    def _extrair_info_produto(self, elemento):
        """Extrai informações detalhadas de um elemento de produto"""
        produto = {}
        
        # Tenta encontrar o elemento de preço específico primeiro
        preco_elementos = elemento.select('[class*="price"]')
        for elem in preco_elementos:
            texto = elem.get_text(strip=True)
            if 'R$' in texto:
                preco = self._limpar_preco(texto)
                if preco:
                    produto['preco'] = preco
                    break
        
        # Se não encontrou preço nos elementos específicos, tenta no texto geral
        if not produto.get('preco'):
            for texto in elemento.stripped_strings:
                if 'R$' in texto and len(texto) < 15:  # Preços geralmente são textos curtos
                    preco = self._limpar_preco(texto)
                    if preco:
                        produto['preco'] = preco
                        break
        
        # Extrai título e URL
        link = elemento.select_one('a[href]')
        if link:
            href = link.get('href')
            if href:
                produto['url'] = urljoin(self.base_url, href) if href.startswith('/') else href
            
            # Tenta extrair título
            titulo = link.get('title', '') or link.get_text(strip=True)
            if titulo:
                produto['titulo'] = titulo

        # Extrai imagem
        img = elemento.select_one('img')
        if img:
            src = img.get('src') or img.get('data-src')
            if src:
                produto['imagem'] = src

        # Extrai avaliação
        rating = elemento.select_one('[class*="rating"]')
        if rating:
            produto['avaliacao'] = rating.get_text(strip=True)

        # Extrai vendas
        vendas = elemento.select_one('[class*="sold"]')
        if vendas:
            produto['vendas'] = vendas.get_text(strip=True)

        # Extrai desconto
        desconto = elemento.select_one('[class*="percent"]')
        if desconto:
            produto['desconto'] = desconto.get_text(strip=True)

        return produto

    def _validar_produtos(self, produtos):
        """Valida e corrige os produtos após a extração"""
        produtos_validados = []
        
        for produto in produtos:
            # Valida preço
            if 'preco' in produto:
                preco_limpo = self._limpar_preco(produto['preco'])
                if preco_limpo:
                    produto['preco'] = preco_limpo
                else:
                    continue  # Remove produto se preço não for válido
            
            produtos_validados.append(produto)
        
        return produtos_validados
    
    def _limpar_produtos(self, produtos):
        """Remove duplicatas, limpa os dados dos produtos e refina o campo título."""
        produtos_limpos = []
        vistos = set()

        for produto in produtos:
            # Limpa espaços em branco
            produto_limpo = {}
            for chave, valor in produto.items():
                if isinstance(valor, str):
                    valor_limpo = re.sub(r'\s+', ' ', valor.strip())
                    if valor_limpo:
                        produto_limpo[chave] = valor_limpo
                else:
                    produto_limpo[chave] = valor

            # Tenta extrair título da URL se não tem título válido
            if produto_limpo.get('url') and (not produto_limpo.get('titulo') or 
                re.match(r'^\d+\s*(vendido|vendidos)', produto_limpo.get('titulo', ''))):
                titulo_url = self._extrair_titulo_da_url(produto_limpo['url'])
                if titulo_url:
                    produto_limpo['titulo'] = titulo_url

            # Refina o campo título para separar desconto, preço, vendas, avaliação
            produto_limpo = self._refinar_titulo(produto_limpo)

            # Verifica se o campo loja não é na verdade um desconto ou avaliação
            if produto_limpo.get('loja'):
                loja = produto_limpo['loja']
                # Se loja parece ser desconto, move para desconto
                if re.match(r'^-?\d+%', loja):
                    produto_limpo['desconto'] = loja
                    del produto_limpo['loja']
                # Se loja parece ser avaliação, move para avaliação
                elif re.match(r'^\d+\.\d+$', loja):
                    produto_limpo['avaliacao'] = loja
                    del produto_limpo['loja']

            # Cria identificador único
            identificador = f"{produto_limpo.get('titulo', '')}-{produto_limpo.get('preco', '')}"

            # Adiciona apenas se não for duplicata e tiver informações mínimas
            if (identificador not in vistos and 
                len(produto_limpo) >= 2 and 
                (produto_limpo.get('titulo') or produto_limpo.get('preco'))):
                vistos.add(identificador)
                produtos_limpos.append(produto_limpo)

        return produtos_limpos
    
    def _calcular_estatisticas(self):
        """Calcula estatísticas dos produtos extraídos"""
        if not self.produtos:
            return
        
        self.estatisticas = {
            'total_produtos': len(self.produtos),
            'com_titulo': sum(1 for p in self.produtos if p.get('titulo')),
            'com_preco': sum(1 for p in self.produtos if p.get('preco')),
            'com_imagem': sum(1 for p in self.produtos if p.get('imagem')),
            'com_url': sum(1 for p in self.produtos if p.get('url')),
            'com_vendas': sum(1 for p in self.produtos if p.get('vendas')),
            'com_avaliacao': sum(1 for p in self.produtos if p.get('avaliacao')),
            'com_loja': sum(1 for p in self.produtos if p.get('loja')),
            'com_desconto': sum(1 for p in self.produtos if p.get('desconto'))
        }
        
        # Análise de preços
        precos = []
        for produto in self.produtos:
            if produto.get('preco'):
                match = re.search(r'[\d.,]+', produto['preco'])
                if match:
                    try:
                        valor = float(match.group().replace('.', '').replace(',', '.'))
                        precos.append(valor)
                    except:
                        pass
        
        if precos:
            self.estatisticas.update({
                'preco_min': min(precos),
                'preco_max': max(precos),
                'preco_medio': sum(precos) / len(precos),
                'precos_analisados': len(precos)
            })
    
    def exibir_relatorio_completo(self):
        """Exibe um relatório completo e bem formatado"""
        print("\n" + "="*80)
        print("                    RELATÓRIO COMPLETO DE EXTRAÇÃO")
        print("="*80)
        
        if not self.produtos:
            print("❌ Nenhum produto foi extraído.")
            print("\n💡 SUGESTÕES:")
            print("1. Verifique se o arquivo HTML contém produtos")
            print("2. Abra o HTML no navegador e use F12 para inspecionar")
            print("3. Procure por padrões como classes CSS específicas")
            print("4. Considere capturar um novo HTML da página")
            return
        
        # Estatísticas principais
        print("\n📊 ESTATÍSTICAS PRINCIPAIS")
        print("-" * 40)
        stats = self.estatisticas
        print(f"Total de produtos extraídos: {stats['total_produtos']:,}")
        print(f"Produtos com título: {stats['com_titulo']:,} ({stats['com_titulo']/stats['total_produtos']*100:.1f}%)")
        print(f"Produtos com preço: {stats['com_preco']:,} ({stats['com_preco']/stats['total_produtos']*100:.1f}%)")
        print(f"Produtos com imagem: {stats['com_imagem']:,} ({stats['com_imagem']/stats['total_produtos']*100:.1f}%)")
        print(f"Produtos com URL: {stats['com_url']:,} ({stats['com_url']/stats['total_produtos']*100:.1f}%)")
        
        # Informações adicionais
        if stats['com_vendas'] > 0:
            print(f"Produtos com info de vendas: {stats['com_vendas']:,}")
        if stats['com_avaliacao'] > 0:
            print(f"Produtos com avaliação: {stats['com_avaliacao']:,}")
        if stats['com_loja'] > 0:
            print(f"Produtos com info da loja: {stats['com_loja']:,}")
        if stats['com_desconto'] > 0:
            print(f"Produtos com desconto: {stats['com_desconto']:,}")
        
        # Análise de preços
        if 'preco_min' in stats:
            print(f"\n💰 ANÁLISE DE PREÇOS")
            print("-" * 40)
            print(f"Menor preço: R$ {stats['preco_min']:.2f}")
            print(f"Maior preço: R$ {stats['preco_max']:.2f}")
            print(f"Preço médio: R$ {stats['preco_medio']:.2f}")
            print(f"Preços analisados: {stats['precos_analisados']:,}")
        
        # Qualidade dos dados
        print(f"\n📈 QUALIDADE DOS DADOS")
        print("-" * 40)
        produtos_completos = sum(1 for p in self.produtos 
                               if len(p) >= 4 and p.get('titulo') and p.get('preco'))
        print(f"Produtos com dados completos: {produtos_completos:,} ({produtos_completos/stats['total_produtos']*100:.1f}%)")
        
        # Exemplos de produtos
        print(f"\n🛍️  EXEMPLOS DE PRODUTOS EXTRAÍDOS")
        print("-" * 40)
        
        # Mostra até 5 produtos mais completos
        produtos_ordenados = sorted(self.produtos, 
                                  key=lambda p: (len(p), bool(p.get('titulo')), bool(p.get('preco'))), 
                                  reverse=True)
        
        for i, produto in enumerate(produtos_ordenados[:5]):
            print(f"\n{i+1}. PRODUTO #{i+1}")
            
            # Verifica se tem título válido
            titulo = produto.get('titulo', 'Título não encontrado')
            # Se o título ainda é vendas, tenta extrair da URL
            if re.match(r'^\d+\s*(vendido|vendidos)', titulo) and produto.get('url'):
                titulo_url = self._extrair_titulo_da_url(produto['url'])
                if titulo_url:
                    titulo = titulo_url
            
            print(f"   📝 Título: {titulo[:80]}..." if len(titulo) > 80 else f"   📝 Título: {titulo}")
            
            # Mostra outros campos disponíveis
            if produto.get('preco'):
                print(f"   💰 Preço: {produto['preco']}")
            else:
                print(f"   💰 Preço: [Não encontrado]")
            
            if produto.get('url'):
                url_display = produto['url'][:60] + '...' if len(produto['url']) > 60 else produto['url']
                print(f"   🔗 URL: {url_display}")
            
            if produto.get('vendas'):
                print(f"   📊 Vendas: {produto['vendas']}")
            
            if produto.get('loja'):
                # Verifica novamente se loja não é desconto ou avaliação
                loja = produto['loja']
                if not re.match(r'^-?\d+%', loja) and not re.match(r'^\d+\.\d+$', loja):
                    print(f"   🏪 Loja: {loja}")
            
            if produto.get('avaliacao'):
                print(f"   ⭐ Avaliação: {produto['avaliacao']}")
            
            if produto.get('desconto'):
                print(f"   🏷️  Desconto: {produto['desconto']}")
            
            if produto.get('frete'):
                print(f"   🚚 {produto['frete']}")
            
            # Se tem imagem, indica
            if produto.get('imagem'):
                print(f"   🖼️  Imagem: Disponível")
        
        # Aviso se muitos produtos estão com títulos incorretos
        produtos_titulo_errado = sum(1 for p in self.produtos 
                                   if p.get('titulo') and re.match(r'^\d+\s*(vendido|vendidos)', p['titulo']))
        
        if produtos_titulo_errado > len(self.produtos) * 0.3:
            print(f"\n⚠️  ATENÇÃO: {produtos_titulo_errado} produtos com títulos incorretos detectados")
            print("💡 Os títulos foram extraídos das URLs quando possível")
            print("   Para melhores resultados, tente capturar o HTML com a página totalmente carregada")
    
    def salvar_resultados(self, formato='todos', pasta_saida=None):
        """Salva os produtos extraídos em múltiplos formatos"""
        if not self.produtos:
            logger.warning("Nenhum produto para salvar")
            return []

        # Ordena produtos por preço
        def extrair_valor(produto):
            preco = produto.get('preco')
            if not preco:
                return float('inf')
            m = re.search(r'[\d.,]+', preco)
            if m:
                try:
                    return float(m.group(0).replace('.', '').replace(',', '.'))
                except:
                    return float('inf')
            return float('inf')
        self.produtos = sorted(self.produtos, key=extrair_valor)

        # Se não passar pasta_saida, salva na pasta atual
        if pasta_saida is None:
            from pathlib import Path
            pasta_saida = Path('.')
        else:
            from pathlib import Path
            pasta_saida = Path(pasta_saida)
            pasta_saida.mkdir(parents=True, exist_ok=True)

        nome_base = os.path.splitext(os.path.basename(self.caminho_arquivo))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        arquivos_salvos = []

        # Salva Excel
        if formato in ['excel', 'todos']:
            try:
                nome_excel = pasta_saida / f'{nome_base}_{timestamp}.xlsx'
                colunas = ['titulo', 'preco', 'desconto', 'avaliacao', 'vendas', 'url', 'imagem']
                import pandas as pd
                df_produtos = pd.DataFrame(self.produtos)
                for col in colunas:
                    if col not in df_produtos.columns:
                        df_produtos[col] = ''
                df_produtos = df_produtos[colunas]
                mapeamento_colunas = {
                    'titulo': 'Título',
                    'preco': 'Preço',
                    'desconto': 'Desconto',
                    'avaliacao': 'Avaliação',
                    'vendas': 'Vendas',
                    'url': 'URL',
                    'imagem': 'Imagem'
                }
                df_produtos = df_produtos.rename(columns=mapeamento_colunas)
                df_produtos.to_excel(str(nome_excel), sheet_name='Produtos Shopee', index=False)

                from openpyxl import load_workbook
                from openpyxl.styles import Alignment
                from openpyxl.utils import get_column_letter

                wb = load_workbook(str(nome_excel))
                ws = wb['Produtos Shopee']

                col_widths = {
                    1: 65,    # Título
                    2: 12,    # Preço
                    3: 12,    # Desconto
                    4: 12,    # Avaliação
                    5: 18,    # Vendas
                    6: 18,    # URL
                    7: 80     # Imagem
                }
                for idx, width in col_widths.items():
                    col_letter = get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = width

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                col_url_idx = list(df_produtos.columns).index('URL') + 1  # +1 pois openpyxl é 1-based

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_url_idx, max_col=col_url_idx):
                    for cell in row:
                        url = str(cell.value).strip() if cell.value else ""
                        if url.startswith('http://') or url.startswith('https://'):
                            cell.hyperlink = url
                            cell.value = "Acessar"
                            cell.style = 'Hyperlink'
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                wb.save(str(nome_excel))

                arquivos_salvos.append(str(nome_excel))
                logger.info(f"Excel salvo: {nome_excel}")
            except Exception as e:
                logger.error(f"Erro ao salvar Excel: {e}")

        # Repita para JSON/CSV se desejar

        print(f"\n💾 ARQUIVOS SALVOS COM SUCESSO")
        print("-" * 40)
        for arquivo in arquivos_salvos:
            tamanho = os.path.getsize(arquivo) / 1024  # KB
            print(f"✓ {arquivo} ({tamanho:.1f} KB)")

        return arquivos_salvos

    def _refinar_titulo(self, produto):
        """
        Refina o campo 'titulo' separando desconto, preço, vendas e avaliação se estiverem juntos.
        """
        import re
        if not produto.get('titulo'):
            return produto

        titulo = produto['titulo']

        # Remove caracteres especiais e espaços extras no início
        titulo = re.sub(r'^[^a-zA-Z0-9]+', '', titulo)

        # Remove "São Paulo" e localização primeiro
        titulo = re.sub(r'\s*São Paulo\b.*$', '', titulo, flags=re.I)
        titulo = re.sub(r'\s+[A-Z]{2}\s*$', '', titulo)

        # Extrai preço (ex: R$49,90)
        preco_match = re.search(r'R\$\s*(\d+[.,]\d{2})', titulo)
        if preco_match and not produto.get('preco'):
            preco = preco_match.group(0)
            produto['preco'] = re.sub(r'R\$\s*(\d+)[.,](\d{2})', r'R$ \1,\2', preco)
            titulo = titulo.replace(preco_match.group(0), '')

        # Extrai desconto (ex: -28% ou 28%)
        desconto_match = re.search(r'-?\d+%', titulo)
        if desconto_match and not produto.get('desconto'):
            desconto = desconto_match.group(0)
            if not desconto.startswith('-'):
                desconto = f"-{desconto}"
            produto['desconto'] = desconto
            titulo = re.sub(r'[\s-]*\d+%[\s-]*', ' ', titulo)

        # Extrai avaliação (ex: 4.9 ou 4,9)
        avaliacao_match = re.search(r'([0-5][.,]\d)', titulo)
        if avaliacao_match and not produto.get('avaliacao'):
            avaliacao = avaliacao_match.group(1).replace(',', '.')
            if 0 <= float(avaliacao) <= 5:
                produto['avaliacao'] = avaliacao
                titulo = titulo.replace(avaliacao_match.group(0), '')

        # Extrai vendas
        vendas_patterns = [
            r'(\d+(?:,\d+)?)\s*mil\s*vendidos?\b',
            r'(\d+(?:\.\d+)?[kK])\s*vendidos?\b',
            r'(\d+)\s*vendidos?\b',
            r'Vendidos:\s*(\d+)',
            r'(\d+)\s*sold\b'
        ]
        for pattern in vendas_patterns:
            vendas_match = re.search(pattern, titulo, re.I)
            if vendas_match and not produto.get('vendas'):
                numero = vendas_match.group(1)
                if 'mil' in pattern.lower() or 'k' in numero.lower():
                    numero = numero.replace('k', '').replace('K', '')
                    try:
                        valor = float(numero.replace(',', '.')) * 1000
                        numero = str(int(valor))
                    except:
                        pass
                produto['vendas'] = f"{numero} vendidos"
                titulo = titulo.replace(vendas_match.group(0), '')

        # Limpeza final do título
        titulo = re.sub(r'\bmil\s*vendidos?\b', '', titulo, flags=re.I)
        titulo = re.sub(r'\s+', ' ', titulo)
        titulo = re.sub(r'^\W+|\W+$', '', titulo)
        titulo = re.sub(r'\s*#', ' #', titulo)
        titulo = re.sub(r'^\d+\s*', '', titulo)
        titulo = re.sub(r'\s+', ' ', titulo).strip()
        titulo = re.sub(r'\s*mil\b', '', titulo, flags=re.I)
        titulo = re.sub(r'\s*vendidos?\b', '', titulo, flags=re.I)
        titulo = re.sub(r'\s+[A-Z]{2}\s*$', '', titulo)
        titulo = re.sub(r'São Paulo\b.*$', '', titulo, flags=re.I)
        titulo = re.sub(r'^[^a-zA-Z0-9]+', '', titulo)
        titulo = re.sub(r'\s+', ' ', titulo).strip()

        produto['titulo'] = titulo

        return produto

def selecionar_arquivo_html():
    """Permite selecionar um arquivo HTML usando o diálogo do Windows"""
    import tkinter as tk
    from tkinter import filedialog
    
    # Cria janela root mas a mantém escondida
    root = tk.Tk()
    root.withdraw()
    
    arquivo = filedialog.askopenfilename(
    title='Selecione o arquivo HTML da Shopee',
    filetypes=[
        ('Arquivos HTML', '*.html;*.htm'),
        ('Todos os arquivos', '*.*')
    ],
    initialdir=r'C:\Users\User\Desktop\Shopee_Copo'  # <-- ALTERE AQUI!
    )
    
    if arquivo:
        tamanho = os.path.getsize(arquivo) / 1024  # KB
        data_mod = datetime.fromtimestamp(os.path.getmtime(arquivo))
        
        print("\n📂 Arquivo selecionado:")
        print(f"   📄 Nome: {os.path.basename(arquivo)}")
        print(f"   📁 Tamanho: {tamanho:.1f} KB")
        print(f"   🕒 Modificado: {data_mod.strftime('%d/%m/%Y %H:%M:%S')}")
        
        return arquivo
    
    return None

if __name__ == "__main__":
    pasta = Path.home() / 'Desktop' / 'prospecção' / 'busca'
    arquivos = [str(f) for f in pasta.glob('*.html')]

    if not arquivos:
        print("❌ Nenhum arquivo HTML encontrado na pasta. Encerrando.")
    else:
        print(f"🔍 Encontrados {len(arquivos)} arquivos HTML para processar.")
        for idx, caminho_arquivo in enumerate(arquivos, 1):
            if parar_execucao:
                print("[INFO] Execução interrompida pelo usuário!")
                break
            print(f"\n===== {idx}/{len(arquivos)} | Arquivo: {os.path.basename(caminho_arquivo)} =====")
            scraper = ShopeeScraper(caminho_arquivo)
            if scraper.carregar_html():
                print("\n🔍 Iniciando extração de produtos da Shopee...")
                produtos = scraper.extrair_produtos_shopee()
                scraper.exibir_relatorio_completo()
                if produtos:
                    scraper.salvar_resultados(formato="excel")
            else:
                print(f"❌ Erro ao carregar o arquivo HTML: {caminho_arquivo}")
        print("\n✅ Processamento de todos os arquivos concluído!")
    
    # Pasta de entrada (onde estão os HTMLs)
    entrada = Path.home() / 'Desktop' / 'prospecção' / 'busca'
    # Pasta de saída desejada
    saida_base = Path.home() / 'Desktop' / 'prospecção' / 'lojas e produtos'

    # Procura todos os HTMLs em todas as subpastas
    arquivos_html = list(entrada.rglob('*.html'))

    print(f"🔍 Encontrados {len(arquivos_html)} arquivos HTML para processar.")

    for caminho_html in arquivos_html:
        if parar_execucao:
            print("[INFO] Execução interrompida pelo usuário!")
            break
        # Subpasta relativa (pode ser '' para raiz, ou o nome da subpasta)
        subpasta_relativa = caminho_html.parent.relative_to(entrada)
        # Cria pasta de saída correspondente
        pasta_saida = saida_base / subpasta_relativa
        pasta_saida.mkdir(parents=True, exist_ok=True)

        print(f"\nProcessando: {caminho_html}")
        print(f"Saída: {pasta_saida}")

        scraper = ShopeeScraper(str(caminho_html))
        if scraper.carregar_html():
            produtos = scraper.extrair_produtos_shopee()
            scraper.exibir_relatorio_completo()
            if produtos:
                arquivos_salvos = scraper.salvar_resultados(formato="excel", pasta_saida=pasta_saida)
                print(f"Arquivos salvos: {arquivos_salvos}")
        else:
            print(f"❌ Erro ao carregar o arquivo HTML: {caminho_html}")