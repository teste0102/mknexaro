#!/usr/bin/env python3
"""
Varredura de preços: Shopee + Mercado Livre → Excel
Uso: python3 varredura_excel.py "tela A15 4g"
"""

import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def buscar_shopee(termo):
    """Busca Shopee via web scraping"""
    resultados = []
    try:
        url = f"https://shopee.com.br/search?keyword={termo.replace(' ', '+')}"
        headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64)"}
        
        resp = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(resp.content, 'html.parser')
        
        # Extrai data-itemid (JSON embutido na página)
        scripts = soup.find_all('script')
        for script in scripts:
            if 'items' in str(script):
                import json
                try:
                    data = json.loads(script.string)
                    for item in data.get('items', [])[:5]:
                        resultados.append({
                            'marketplace': 'SHOPEE',
                            'produto': item.get('name', 'N/A'),
                            'preco': f"R$ {item.get('price', 0) / 100000:.2f}",
                            'vendedor': item.get('shop', {}).get('name', 'N/A'),
                            'url': f"https://shopee.com.br/search?keyword={termo}"
                        })
                except:
                    pass
    except Exception as e:
        print(f"Erro Shopee: {e}")
    
    return resultados

def buscar_mercadolivre(termo):
    """Busca Mercado Livre via web scraping"""
    resultados = []
    try:
        url = f"https://www.mercadolivre.com.br/search?q={termo.replace(' ', '+')}"
        headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64)"}
        
        resp = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(resp.content, 'html.parser')
        
        # Encontra produtos (classes do ML)
        produtos = soup.find_all('div', {'class': 'andes-card'})[:5]
        
        for prod in produtos:
            try:
                nome = prod.find('a', {'class': 'poly-component__title'})
                preco = prod.find('span', {'class': 'price-tag-fraction'})
                vendedor = prod.find('p', {'class': 'poly-component__seller'})
                
                if nome and preco:
                    resultados.append({
                        'marketplace': 'MERCADO LIVRE',
                        'produto': nome.text.strip(),
                        'preco': f"R$ {preco.text.strip()}",
                        'vendedor': vendedor.text.strip() if vendedor else 'N/A',
                        'url': f"https://www.mercadolivre.com.br/search?q={termo}"
                    })
            except:
                pass
    except Exception as e:
        print(f"Erro ML: {e}")
    
    return resultados

def salvar_excel(dados, termo):
    """Salva dados em Excel com formatação"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Varredura"
    
    # Header
    headers = ['Marketplace', 'Produto', 'Preço', 'Vendedor', 'URL', 'Status']
    ws.append(headers)
    
    # Formatação header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adiciona dados
    for row in dados:
        ws.append([
            row['marketplace'],
            row['produto'],
            row['preco'],
            row['vendedor'],
            row['url'],
            'VALIDAR'  # Coluna para você marcar após revisar
        ])
    
    # Ajusta colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 12
    
    # Salva
    filename = f"varredura_{termo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    return filename

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 varredura_excel.py 'termo'")
        sys.exit(1)
    
    termo = sys.argv[1]
    
    print(f"🔍 Buscando '{termo}' em Shopee e Mercado Livre...")
    print("Isso pode levar alguns segundos...\n")
    
    dados = []
    dados.extend(buscar_shopee(termo))
    dados.extend(buscar_mercadolivre(termo))
    
    if dados:
        arquivo = salvar_excel(dados, termo)
        print(f"✅ Arquivo criado: {arquivo}")
        print(f"📊 {len(dados)} produtos encontrados")
        print("\n💡 Abra o arquivo, revise os preços e marque como VALIDADO")
    else:
        print("❌ Nenhum produto encontrado")
