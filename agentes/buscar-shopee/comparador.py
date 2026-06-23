import pandas as pd
import os
import glob
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from tkinter import messagebox

# Você pode importar customtkinter, mas só para usar os messagebox mesmo
# Se quiser usar messagebox CTk: 
#import customtkinter as ctk

cores_planilhas = [
    "FFFF33", "99FF00", "FF3300", "FF00FF", "ff9900",
    "6b705c", "0000FF", "CC00FF", "66FF00", "33ff33"
]
cor_preco_verde = "07ED16"

def comparar_produtos_excels(termo_busca, lista_pastas, pasta_saida, modo_exibicao="todos"):
    """
    termo_busca: string do termo procurado
    lista_pastas: lista de caminhos (strings) das pastas com excels
    pasta_saida: pasta onde salvar o excel final consolidado
    modo_exibicao: "1", "2", "todos"  (mais barato, top 5, todos)
    """
    resultados = []
    
    # Formatando a data de busca para garantir que esteja no mesmo formato que a extraída dos arquivos.
    try:
        termo_busca = termo_busca.strip().lower()  # Garantir que o termo seja em minúsculas
    except Exception as e:
        print(f"Erro ao processar termo de busca: {e}")
    
    for idx, pasta in enumerate(lista_pastas):
        arquivos = sorted(glob.glob(os.path.join(pasta, '*.xlsx')), reverse=True)
        
        for arquivo in arquivos:
            try:
                # Lê o arquivo Excel
                df = ler_excel_com_links(arquivo)
                
                # Verifique se as colunas 'Título' e 'Preço' existem
                if 'Título' not in df.columns or 'Preço' not in df.columns:
                    continue
                
                # Filtragem do termo de busca no título (insensível a maiúsculas e minúsculas)
                resultado = df[df['Título'].str.contains(termo_busca, case=False, na=False)].copy()
                
                if not resultado.empty:
                    # Adicionando informações extras
                    resultado['Fonte'] = os.path.basename(arquivo)
                    resultado['Pasta'] = os.path.basename(pasta)
                    resultado['OrigemIdx'] = idx
                    
                    # Processando os preços para garantir que a formatação seja consistente
                    resultado['Preço'] = (
                        resultado['Preço']
                        .astype(str)
                        .str.replace('R$', '', regex=False)  # Remove o símbolo de moeda
                        .str.replace(',', '.', regex=False)  # Substitui vírgulas por ponto
                        .str.extract(r'([\d.]+)')[0]  # Extrai apenas números e ponto
                        .astype(float)  # Converte para float
                    )
                    resultados.append(resultado)
            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}")

    if not resultados:
        messagebox.showinfo("Resultado", f"Nenhum produto '{termo_busca}' encontrado nas planilhas selecionadas.")
        return None

    # Concatenando todos os resultados
    df_final = pd.concat(resultados, ignore_index=True)
    df_final = df_final.sort_values(by=['Preço', 'Fonte'], ascending=[True, True])

    # Filtro por modo de exibição
    if modo_exibicao == "1":
        df_exibir = df_final.head(1)  # Exibir o mais barato
    elif modo_exibicao == "2":
        df_exibir = df_final.head(5)  # Exibir top 5 mais baratos
    else:
        df_exibir = df_final  # Exibir todos os resultados

    # Gerar o nome do arquivo de saída
    nome_produto_limpo = ''.join(c for c in termo_busca if c.isalnum() or c in [' ', '_', '-']).strip()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_arquivo = f"{timestamp}_{nome_produto_limpo.replace(' ', '_')}.xlsx"
    caminho_exportacao = os.path.join(pasta_saida, nome_arquivo)
    os.makedirs(pasta_saida, exist_ok=True)

    # Colunas que vão ser exportadas
    colunas_exportar = [
        'Título', 'Preço', 'Desconto', 'Avaliação', 'Vendas',
        'URL', 'Fonte', 'Pasta', 'Imagem'
    ]
    
    for col in colunas_exportar:
        if col not in df_final.columns:
            df_final[col] = ""

    df_export = df_final[colunas_exportar]

    # Salvando o Excel com a formatação desejada
    with pd.ExcelWriter(caminho_exportacao, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Produtos', index=False)
        ws = writer.sheets['Produtos']

        # Cabeçalho com formatação
        header_font = Font(name='Calibri', bold=True, size=11)
        header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths
        from openpyxl.utils import get_column_letter
        col_widths = {
            1: 65,    # Título
            2: 12,    # Preço
            3: 12,    # Desconto
            4: 12,    # Avaliação
            5: 18,    # Vendas
            6: 18,    # URL
            7: 30,    # Fonte (Arquivo)
            8: 80,    # Pasta
            9: 80     # Imagem
        }
        for idx, width in col_widths.items():
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

        # Center alignment and formatting
        centralizar_cols = ['Preço', 'Desconto', 'Avaliação', 'URL', 'Pasta', 'Vendas', 'Fonte']
        col_idx_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

        # Price in green
        preco_col = col_idx_map.get('Preço')
        if preco_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[preco_col-1]
                cell.font = Font(color="07ED16")  # Green
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Center other columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_name in centralizar_cols:
                col_idx = col_idx_map.get(col_name)
                if col_idx and col_name != 'Preço':  # Price already centered above
                    cell = row[col_idx-1]
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # URL as clickable hyperlink
        url_col = col_idx_map.get('URL')
        if url_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[url_col-1]
                url = str(cell.value).strip() if cell.value else ""
                if url.startswith("http://") or url.startswith("https://"):
                    cell.value = "Acessar"
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Colorir as linhas de acordo com a origem e o preço
        for row in range(2, ws.max_row + 1):
            origem_idx = ws.cell(row=row, column=colunas_exportar.index('OrigemIdx') + 1).value
            cor_fundo = cores_planilhas[origem_idx % len(cores_planilhas)] if origem_idx is not None else "FFFFFF"
            fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type='solid')
            for col in range(1, len(colunas_exportar) + 1):
                cell = ws.cell(row=row, column=col)
                if col == colunas_exportar.index('Preço') + 1:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    cell.font = Font(color=cor_preco_verde)
                else:
                    cell.fill = fill

        ws.delete_cols(len(colunas_exportar))  # Apaga as colunas extras
        ws.freeze_panes = 'A2'  # Congela a primeira linha (cabeçalho)

    messagebox.showinfo("Exportado", f"Produtos exportados em:\n{caminho_exportacao}")
    return caminho_exportacao

def ler_excel_com_links(path):
    df = pd.read_excel(path)
    wb = load_workbook(path)
    ws = wb.active
    url_col_idx = None
    # Procura a coluna 'URL' no cabeçalho
    for i, cell in enumerate(ws[1]):
        if cell.value and str(cell.value).strip().lower() == 'url':
            url_col_idx = i
            break
    if url_col_idx is not None:
        urls_reais = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[url_col_idx]
            # Pega o hyperlink real se existir, senão pega o valor da célula
            if cell.hyperlink:
                urls_reais.append(cell.hyperlink.target)
            else:
                urls_reais.append(cell.value)
        df['URL'] = urls_reais
    return df
